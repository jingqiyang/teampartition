from datetime import datetime, timedelta
from math import ceil
from openpyxl import load_workbook
from statistics import stdev
import sys

from partition_constants import *

# current time
now = datetime.now()
year = str(now.year)


###### LOAD DATA FUNCTIONS ######

"""
get dict player data from command line input or default test excel workbook.
"""
def getPlayers():
    wb_file = sys.argv[1] if len(sys.argv) > 1 else TEST_WB

    players_wb = load_workbook(wb_file, data_only=True)
    players_ws = players_wb[year]
    players = getData(players_ws)

    mergeConflicts(players)
    return players


"""
get data from worksheet as dict with first column values as keys, and top row as fields of each element.
"""
def getData(ws):
    data = {}
    cols = [cell.value for cell in ws[1]]

    r = 2
    key = ws.cell(column=1, row=r).value

    while key:
        element = {}
        for c in range(0, len(cols)):
            col = cols[c]
            if col not in element:
                element[col] = ws.cell(column=c + 1, row=r).value

        if key in data:
            print("warning: duplicate element " + key + ", skipping duplicate")
        else:
            data[key] = element

        r += 1
        key = ws.cell(column=1, row=r).value

    return data


"""
combine player conflicts into 1 set and check for one-sided conflicts.
"""
def mergeConflicts(players):
    # initialize empty set of conflicts for each player
    for p in players:
        players[p][CONFLICTS] = set()

    for p in players:
        player = players[p]

        # construct excel column "Conflict 1", "Conflict 2", etc
        i = 1
        conflict_key = CONFLICT + " " + str(i)

        while conflict_key in player and player[conflict_key]:
            conflict_p = player[conflict_key]
            player[CONFLICTS].add(conflict_p)

            # make it a mutual conflict
            players[conflict_p][CONFLICTS].add(p)

            i += 1
            conflict_key = CONFLICT + " " + str(i)



###### INTERPRETING PLAYER DATA FUNCTIONS ######

def isSenior(player):
    return player[SENIOR] == YES

def isSetter(player):
    return player[SETTER] == YES


"""
convert yankee rating to numerical score.
"""
def getScore(rating):
    letter = rating[0].upper()
    if letter < "A" or letter > "D":
        print("warning: invalid yankee letter rating")
        return None

    # use ascii value of letter with D as min score of 0
    score = 68 - ord(letter)

    if rating.endswith("-"):
        return score - 1/3
    elif rating.endswith("+"):
        return score + 1/3
    return score


"""
get overall player score as a sum of offense/setting and defense, adjusted by if they're a senior.
"""
def getOverallScore(player):
    defense_rating = player[SETTING] if isSetter(player) else player[DEFENSE]
    offense_rating = player[OFFENSE]
    senior_modifier = -1/3 if isSenior(player) else 0

    player[OVERALL] = getScore(offense_rating) + getScore(defense_rating) + senior_modifier


"""
get float score value as string.
"""
def getScoreString(score):
    return str(round(score, 2))


"""
print all scores of players in gender/position/age groups.
"""
def printScores(player_groups, players):
    for player_keys in player_groups:
        for p in player_keys:
            print(p + ": " + getScoreString(players[p][OVERALL]))
        print()



###### DATA MANIPULATION FUNCTIONS ######

"""
partition a sub list from an original list based on a lambda function returning true for each element.
the elements of the sub list are removed from the original list.
"""
def splitList(original_list, f):
    sub_list = []
    for i in range(len(original_list) - 1, -1, -1):
        if f(original_list[i]):
            element = original_list.pop(i)
            sub_list.insert(0, element)

    return sub_list, original_list



###### TEAM LIST MANIPULAITON FUNCTIONS ######

"""
create list of empty team lists.
"""
def initTeams(num_players):
    return [[] for x in range(0, ceil(num_players / MAX_TEAM_SIZE))]


"""
create list of empty team sets.
"""
def initTeamSets(num_players):
    return [set() for x in range(0, ceil(num_players / MAX_TEAM_SIZE))]


"""
assign players to teams in a snaking pattern. returns unassigned conflict players.
"""
def assignTeams(player_keys, teams, players, conflicts):
    i = 0
    sortTeams(teams, players, reverse=True)

    for p in player_keys:
        assigned = False
        started = False
        init_i = i

        while not assigned and (not started or i != init_i):
            started = True
            team = teams[i]

            # don't add to team with more players until all teams reach same number of players
            if okToAdd(team, players[p], teams, players):
                team.append(p)
                assigned = True

            i += 1

            # turn the snake around
            if i == len(teams):
                teams.reverse()
                i = 0

        # track conflict player to assign later
        if not assigned:
            conflicts.append(p)


"""
sort teams by average player score.
"""
def sortTeams(teams, players, reverse=False):
    teams.sort(key=lambda t: getTeamScore(t, players), reverse=reverse)


"""
determine if a player can be added to a team.
"""
def okToAdd(team, player, teams, players):
    # check for acceptable team size
    if len(team) < MAX_TEAM_SIZE - 1 or all(map(lambda t: len(t) >= MAX_TEAM_SIZE - 1, teams)):
        return noConflict(player, team)

    return False


"""
check if a player has no conflict with existing members of a team.
"""
def noConflict(player, team):
    team = set(team)
    for p in player[CONFLICTS]:
        if p in team:
            return False

    return True


"""
add unassigned conflict players and make swaps until all teams have enough players.
"""
def assignConflicts(teams, players, conflicts):
    for i in range(len(conflicts) - 1, -1, -1):
        p = conflicts[i]
        assigned = False

        # assign conflicts to teams that aren't full where possible
        for team in teams:
            if len(team) < MAX_TEAM_SIZE and noConflict(players[p], team):
                team.append(p)
                del conflicts[i]
                assigned = True
                break

    # TODO: continually make swaps to assign remaining conflict players

    if len(conflicts) > 0:
        print("unassigned conflicts: " + ", ".join(conflicts) + "\n")


"""
make swaps to decrease team score standard deviation.
"""
def finalSwaps(teams, players):
    sortTeams(teams, players)
    std = getTeamStd(teams, players)

    # swap players at same index across teams with major to minor score difference
    for i in range(0, len(teams[0])):
        for j in range(0, int(len(teams) / 2)):
            # try swap with adjacent indices
            for k in range(i - 1, i + 2):
                swap(teams[j], i, teams[-1 - j], k, players)
                new_std = getTeamStd(teams, players)

                if new_std < std:
                    std = new_std
                else:
                    # swap back if no improvement
                    swap(teams[j], i, teams[-1 - j], k, players)

        sortTeams(teams, players)


"""
get standard deviation of all teams' scores.
"""
def getTeamStd(teams, players):
    return stdev(map(lambda t: getTeamScore(t, players), teams))


"""
swap player from team a at index a with player from team b at index b.
"""
def swap(team_a, a, team_b, b, players):
    if not playersMatch(team_a, a, team_b, b, players):
        return

    player_a = team_a[a]
    player_b = team_b[b]

    if noConflict(players[player_a], team_b) and noConflict(players[player_b], team_a):
        team_a[a] = player_b
        team_b[b] = player_a


"""
determine if players have matching characteristics to allow a swap.
"""
def playersMatch(team_a, a, team_b, b, players):
    # no swap if index out of bounds
    if len(team_a) <= a or len(team_b) <= b:
        return False

    player_a = team_a[a]
    player_b = team_b[b]

    for field in [GENDER, SENIOR, SETTER]:
        if players[player_a][field] != players[player_b][field]:
            return False

    return True


"""
get average overall score of players of a team.
"""
def getTeamScore(team, players):
    if len(team) == 0:
        return 0
    return sum(map(lambda p: players[p][OVERALL], team)) / len(team)


"""
print sorted players and player scores, followed by the average overall score of each team.
"""
def printTeamScores(teams, players):
    for team in teams:
        if len(team) == 0:
            print("warning: empty team")
            continue

        for p in team:
            print(p + " " + getScoreString(players[p][OVERALL]))
        print("team score: " + getScoreString(getTeamScore(team, players)) + " (size " + str(len(team)) + ")\n")



###### OBSELETE FUNCTIONS ######

# """
# invert a list of teams so the outer values are on the inside and vice versa.
# """
# def invert(teams):
#     midpoint = int(len(teams) / 2)

#     return list(reversed(teams[:midpoint])) + list(reversed(teams[midpoint:]))


# """
# recursively assign players to teams by first assigning the start and end of the player list.
# """
# def assignPlayers(player_keys, teams, players):
#     num_teams = len(teams)
#     num_players = len(player_keys)

#     # base case and repeat every recursive iteration: assign "outer rows" of players
#     if num_players < num_teams:
#         sortTeams(teams, players, reverse=True)
#     assignSequential(player_keys[:num_teams], teams, players)
#     assignSequential(list(reversed(player_keys[num_teams:])), teams, players)

#     # assign inner rows
#     if num_players > num_teams * 2:
#         assignPlayers(player_keys[num_teams:-num_teams], teams, players)


# """
# assign a list of players to the teams in sequential order.
# """
# def assignSequential(player_keys, teams, players):
#     i = 0

#     for p in player_keys:
#         added_player = False

#         while not added_player:
#             if i == len(teams):
#                 i = 0

#             team = teams[i]

#             if len(team) < MAX_TEAM_SIZE - 1 or all(map(lambda t: len(t) >= MAX_TEAM_SIZE - 1, teams)):
#                 team.append(p)
#                 added_player = True

#             i += 1
